# 这个文档会记录量化测试，每一人口的情况下，进行1000场战斗，随机棋子/限制棋子星级后随机棋子。
# 记录棋子胜率，棋子平均伤害/单局最高伤害，棋子平均承受伤害/单局最高承受伤害，棋子技能施放频率，
from battle import *
from copy import deepcopy
from json import dumps
from numpy.random import randint
import openpyxl 
from util import *

everythingWeNeedToKnowAboutAChess = {'chessName': '',
                                     'totalMatch': 0, 
                                     'winCount':0, 
                                     'maxDamageInOneMatch': 0, # 单局造成最高伤害
                                     'totalDamage': 0,# 总伤害
                                     'totalAttackDamage': 0, # 总攻击伤害
                                     'totalSpellDamage': 0, # 总技能伤害
                                     'maxDamageReceivedInOneMatch':0, # 单局受到最高伤害
                                     'totalDamageReceived':0, # 总承受伤害
                                     'totalSpellCasted': 0, # 技能施放总数
                                     'totalMatchLength': 0, # 总战斗时长
                                     'longestMatchLength': 0, # 最长战斗时长
                                     'shortestMatchLength': 100000, # 最短战斗时长
                                     'totalLivingLength': 0, # 总在场时长
                                     'winRate': None, # 总胜率
                                     'averageMatchLength': None,
                                     'averageDamagePerMatch': None,
                                     'averageDamageReceivedPerMatch': None,
                                     
                                     # 和那些棋子一起上场的时候胜率高
                                     }
chessBattleResult_dict: dict[int,dict[str,float]] = {} 
for id, chessChineseName in chessName_dict_inEng.items():
    chessBattleResult_dict[id] = deepcopy(everythingWeNeedToKnowAboutAChess)
    chessBattleResult_dict[id]['chessName'] = chessChineseName
    
# 在没做棋子判定随机的情况下，哪个队伍更有可能获胜。
redTeamWinRate = 0.5
blueTeamWinRate = 1-redTeamWinRate

turnRestrains = {1: {1:1, 2:0,3:0,4:0},  # [1, 0, 0, 0]
                  2: {1:3, 2:1,3:0,4:0},  # [0.5, 0.5, 0, 0]
                  3: {1:6, 2:3,3:1,4:0},  # [0.3, 0.5, 0.2, 0]
                  4: {1:9, 2:4,3:2,4:1},  # [0.15, 0.4, 0.4, 0.05] # 如果有四星就不能满人口
                  5: {1:12,2:6,3:3,4:1},  # [0.1, 0.25, 0.4, 0.15] # 必有一个三星
                  6: {1:16,2:8,3:4,4:2},  # [0.1, 0.25, 0.5, 0.15] # 必有一个四星
                  7: {1:21,2:10,3:5,4:2}, # [0.05, 0.25, 0.4, 0.3]
                  8: {1:28,2:14,3:7,4:3}, # [0.05, 0.25, 0.4, 0.3]
                  9: {1:34,2:17,3:8,4:4}, # [0.05, 0.25, 0.35, 0.35]
                  10: {1:40,2:20,3:10,4:5}}
turnChessPossibility = {
                  1:[1, 0, 0, 0],            # 必有一个一星
                  2: [0.5, 0.5, 0, 0],       # 必有一个2星
                  3: [0.3, 0.5, 0.2, 0],     # 纯随机
                  4: [0.15, 0.4, 0.4, 0.05], # 如果有四星就不能满人口
                  5: [0.1, 0.4, 0.4, 0.1], # 必有一个三星
                  6: [0.1, 0.6, 0.3, 0], # 必有一个四星
                  7: [0.00, 0.45, 0.45, 0.1], # 必有一个四星，一个三星
                  8: [0.00, 0.35, 0.45, 0.2], # 必有一个四星, 两个三星
                  9: [0.0, 0.25, 0.45, 0.35],
                  10: [0.0, 0.25, 0.35, 0.45]}
def chess_level(pop, turn) -> list[int]:
    """根据人口和回合，返回一个以棋子星级为值的list

    Args:
        pop (_type_): _description_
        turn (_type_): _description_

    Returns:
        list[int]: _description_
    """
    possibilities = deepcopy(turnChessPossibility[turn])
    chessLevels = []
    for i in range(pop):
        rand = random()
        # print("pop= ",i, " randomValue: ",rand)
        if i == 0:
            # 根据回合数必定获得这个星级棋子
            if turn == 1:
                chessLevels.append(1)
                continue
            elif turn == 2:
                chessLevels.append(2)
                continue
            elif turn == 4 and rand>0.6: 
                chessLevels.append(3)
                continue
            elif turn == 5:
                chessLevels.append(3)
                continue
            elif turn == 6:
                chessLevels.append(4)
                continue
            elif turn == 7:
                chessLevels.append(4)
                continue
            elif turn == 8:
                chessLevels.append(4)
                continue
            elif turn == 9:
                chessLevels.append(4)
                continue
            elif turn == 10:
                chessLevels.append(4)
                continue
        elif i == 1:
            if turn == 9:
                chessLevels.append(4)
                continue
            elif turn == 10:
                chessLevels.append(4)
                continue
        # 根据概率随机相应棋子
        if rand <= possibilities[0]:
            chessLevels.append(1)
        elif rand > possibilities[0] and \
                rand <= possibilities[0]+ possibilities[1]:
            chessLevels.append(2)
        elif rand > possibilities[0]+ possibilities[1] and \
                rand <=  possibilities[0]+ possibilities[1]+ possibilities[2]:
            chessLevels.append(3)
        else:
            chessLevels.append(4)
    return chessLevels    
    
turnPopulationRestriction = {
                  1: [1], # 括号中是每回合可能出现的人口
                  2: [2],
                  3: [2],
                  4: [3],
                  5: [3,4],
                  6: [3,4],
                  7: [4,5],
                  8: [4,5], 
                  9: [4,5,6],
                  10: [4,5,6]}

# 每一星棋子选择
star1 = [1,2,3]
star2 = [4,5,6,7,18,9]
star3 = [10,11,12,13,14,15,16,17,30,19,21,22] # 20 还没加进去
star4 = [23,24,25,26,27,28]
        
        
def get_random_chess_with_level(level:int):
    if level > 4:
        return None
    if level == 1:
        chessID = star1[randint(0,len(star1))]
    elif level == 2:
        chessID = star2[randint(0,len(star2))]
    elif level == 3:
        chessID = star3[randint(0,len(star3))]
    elif level == 4:
        chessID = star4[randint(0,len(star4))]
    chess = chess_dict[chessID]
    return  chess

def get_random_chesses_with_pop(population:int, turn: int, team: int):
    chessLevels = chess_level(pop=population,turn=turn)
    rowRange = []
    colRange = [0,1,2,3,4]
    positions = []
    if team == 0:
        rowRange =[0,1,2]
    else:
        rowRange =[3,4,5]
    i = 1
    while i <= population:
        position = [rowRange[randint(0,len(rowRange))], colRange[randint(0,len(colRange))]]
        if position not in positions:
            positions.append(position)
            i += 1
    
    chesses = [] # a list of initiated chess
    for j in range(population):
        chessLevel  = chessLevels[j]
        # print(position)
        chessPosition = positions[j]
        chess = get_random_chess_with_level(chessLevel)
        
        chesses.append(chess(position=chessPosition))
    return chesses
        
    # random chess positions 
    
    
def random_battle(turn: int, n):
    """随机 n 个单局游戏 并且记录战斗信息
    """
    redTeamWins = 0
    blueTeamWins = 0
    for i in range(n): # 循环n次
        populationRestriction = turnPopulationRestriction[turn]
        redPop = populationRestriction[randint(0,len(populationRestriction))]
        bluePop = populationRestriction[randint(0,len(populationRestriction))]
        redTeam: list[chessInterface] = get_random_chesses_with_pop(population = redPop, turn = turn, team = 0)
        blueTeam: list[chessInterface] = get_random_chesses_with_pop(population = bluePop, turn = turn, team = 1)

        newBattle = battle()
        newBattle.board_print()
        newBattle.addRedTeam(redTeam)
        newBattle.addBlueTeam(blueTeam)
        wonTeam, matchLength = newBattle.battle_with_skills_test()
        
        for (uniqueID, chess) in newBattle.allChessDict.items():
            chessBattleResult_dict[chess.id]['totalMatch'] += 1
            if wonTeam == chess.team:
                chessBattleResult_dict[chess.id]['winCount'] += 1
            # 记录伤害
            if chess.totalDamage > chessBattleResult_dict[chess.id]['maxDamageInOneMatch']:
                chessBattleResult_dict[chess.id]['maxDamageInOneMatch'] = chess.totalDamage
            chessBattleResult_dict[chess.id]['totalDamage'] += chess.totalDamage
            
            chessBattleResult_dict[chess.id]['totalAttackDamage'] += chess.totalAttackDamage
            chessBattleResult_dict[chess.id]['totalSpellDamage'] += (chess.totalDamage -chess.totalAttackDamage)
            # 记录承伤
            if chess.totalDamageReceived > chessBattleResult_dict[chess.id]['maxDamageReceivedInOneMatch']:
                chessBattleResult_dict[chess.id]['maxDamageReceivedInOneMatch'] = chess.totalDamageReceived
            chessBattleResult_dict[chess.id]['totalDamageReceived'] += chess.totalDamageReceived
            # 记录对局信息
            chessBattleResult_dict[chess.id]['totalMatchLength'] += matchLength
            if matchLength > chessBattleResult_dict[chess.id]['longestMatchLength']:
                chessBattleResult_dict[chess.id]['longestMatchLength'] = matchLength
            elif matchLength < chessBattleResult_dict[chess.id]['shortestMatchLength']:
                chessBattleResult_dict[chess.id]['shortestMatchLength'] = 1
            if chess.isDead:
                chessBattleResult_dict[chess.id]['totalLivingLength'] += chess.deathTime
                
                                    #  'totalSpellCasted': 0, # 技能施放总数
        if wonTeam == 0:
            redTeamWins += 1
        else:
            blueTeamWins += 1
        print("Battle: ", i," finished")
    
            
    
def main():
    for turn in range(1,11):
        # random a population for that guy
        print("\nturn: ", turn)
        random_battle(turn = turn, n = 200)
    # print(chessBattleResult_dict)

    #  'averageDamageReceivedPerMatch': None,
    for cid, chess in chessBattleResult_dict.items():
        totalBattles = chessBattleResult_dict[cid]['totalMatch']
        if totalBattles == 0:
            totalBattles += 1
        wonBattles = chessBattleResult_dict[cid]['winCount']
        chessBattleResult_dict[cid]['winRate'] = round(wonBattles/totalBattles,3) * 100
        
        totalMatchLen = chessBattleResult_dict[cid]['totalMatchLength']
        chessBattleResult_dict[cid]['averageMatchLength'] = round(totalMatchLen/totalBattles,3)
        
        totalDamage = chessBattleResult_dict[cid]['totalDamage']
        chessBattleResult_dict[cid]['averageDamagePerMatch'] = round(totalDamage/totalBattles,3)
        
        totalDamageReceived = chessBattleResult_dict[cid]['totalDamageReceived']
        chessBattleResult_dict[cid]['averageDamageReceivedPerMatch'] = round(totalDamageReceived/totalBattles,3)


    # 保存结果到excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Battle Data"
    col = 1
    for attr in everythingWeNeedToKnowAboutAChess.keys():
        ws.cell(row=1, column=col, value = attr)
        col += 1
    row = 2
    for chessID, chessData in chessBattleResult_dict.items():
        col = 1
        for value in chessData.values():
            ws.cell(row=row, column=col, value = value)
            col += 1
        row += 1 
    wb.save('battleData.xlsx')
    # save to json
    # with open("battle_results.json", 'w') as f:
    #     f.write(dumps(chessBattleResult_dict))
    
    
if __name__ == '__main__':
    main()